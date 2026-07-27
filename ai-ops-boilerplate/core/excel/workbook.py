"""Safe workbook I/O.

The discipline this module enforces:

* **One open, one close.** :func:`open_workbook` is a context manager that
  guarantees ``wb.close()`` even on exception. Never call ``load_workbook``
  directly in automation code.
* **Atomic save.** :func:`atomic_save` writes to a sibling temp file and then
  ``os.replace``s it over the target, so an interrupted save can never leave a
  half-written (corrupt) workbook or clobber the original.
* **Read once.** :func:`read_table` opens read-only, pulls a sheet into a
  pandas DataFrame, and closes — the input is never held open across work.

These three primitives replace the open→save→reopen cycles and leaked handles
that previously corrupted output files.
"""

from __future__ import annotations

import os
import time
from collections.abc import Iterator
from contextlib import contextmanager
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.workbook import Workbook

# Extensions openpyxl recognises. WorkingMemory stores files without an
# extension, so for anything else we open via a binary handle to bypass
# openpyxl's filename-suffix check.
_KNOWN_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xltm"}


@contextmanager
def open_workbook(
    path: str | Path,
    *,
    read_only: bool = False,
    data_only: bool = False,
    keep_vba: bool = False,
) -> Iterator[Workbook]:
    """Open a workbook and guarantee it is closed.

    ``data_only=True`` reads the values Excel last cached for formula cells.
    Note the framework never relies on this to obtain a value it computed —
    authoritative values live in the data model, not in re-read formula cells.
    """
    path = Path(path)
    handle = None
    if path.suffix.lower() in _KNOWN_SUFFIXES:
        wb = load_workbook(
            path, read_only=read_only, data_only=data_only, keep_vba=keep_vba
        )
    else:
        # Open via handle so openpyxl skips its extension check.
        handle = open(path, "rb")
        wb = load_workbook(
            handle, read_only=read_only, data_only=data_only, keep_vba=keep_vba
        )
    try:
        yield wb
    finally:
        wb.close()
        if handle is not None:
            handle.close()


def atomic_save(
    wb: Workbook, path: str | Path, *, replace_attempts: int = 5, replace_delay: float = 0.1
) -> Path:
    """Save ``wb`` to ``path`` atomically (temp file + ``os.replace``).

    Returns the destination path. The temp file is removed on failure so no
    ``.tmp`` litter is left behind.

    On Windows ``os.replace`` can transiently raise ``PermissionError`` when the
    destination is momentarily locked (antivirus, a not-yet-released handle, the
    file open in Excel). The replace is retried a few times with a short backoff
    before giving up.
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_name(f".{path.name}.tmp")
    try:
        wb.save(tmp)
        for attempt in range(replace_attempts):
            try:
                os.replace(tmp, path)  # atomic on the same volume (incl. Windows)
                break
            except PermissionError:
                if attempt == replace_attempts - 1:
                    raise
                time.sleep(replace_delay)
    except BaseException:
        if tmp.exists():
            try:
                tmp.unlink()
            except OSError:
                pass
        raise
    return path


def read_table(
    path: str | Path,
    *,
    sheet: str | None = None,
    header_row: int = 1,
    start_row: int | None = None,
    data_only: bool = True,
) -> pd.DataFrame:
    """Read a sheet into a DataFrame, opening and closing the file once.

    ``header_row`` is the 1-based row holding column headers; data is read from
    ``start_row`` (default ``header_row + 1``) onward. Columns are keyed by the
    header text (blank headers become ``col_<n>``).
    """
    if start_row is None:
        start_row = header_row + 1
    with open_workbook(path, read_only=True, data_only=data_only) as wb:
        ws = wb[sheet] if sheet else wb.active
        rows = ws.iter_rows(values_only=True)
        # Advance to the header row.
        header: list[str] = []
        for idx, row in enumerate(rows, start=1):
            if idx == header_row:
                header = [
                    str(h) if h is not None and str(h).strip() != "" else f"col_{i}"
                    for i, h in enumerate(row, start=1)
                ]
                break
        if not header:
            return pd.DataFrame()
        # Skip any rows between header and start_row.
        skip = start_row - header_row - 1
        records: list[dict[str, Any]] = []
        for offset, row in enumerate(rows):
            if offset < skip:
                continue
            records.append({header[i]: v for i, v in enumerate(row) if i < len(header)})
    return pd.DataFrame.from_records(records, columns=header)
