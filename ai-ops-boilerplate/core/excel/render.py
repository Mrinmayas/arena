"""Render SheetModels to a workbook in a single write pass.

Discipline enforced here:

* **Preserve-by-default.** The renderer touches only the sheets named by the
  models passed in. Unrelated sheets (e.g. a ``Sumifs`` sheet) survive untouched
  — the structural fix for the old "delete a sheet to clean up" data loss.
* **Bounded replace, not blind clear.** Each sheet's data region is overwritten
  with exactly the model's rows; residual old rows are cleared; blocks below the
  data region (``PreservedRegion``s) are relocated by the row delta.
* **One write open.** The write happens in a single ``open → apply → atomic_save
  → close`` pass. Freeze-value preservation adds one *read-only* capture pass
  beforehand (a disciplined read, not the old reopen-to-write cycle).
* **Integrity guard.** The renderer raises :class:`WorkbookIntegrityError` if a
  render would drop a sheet. Preserved regions are relocated below the data
  region collision-free by construction (the data↔region padding is preserved).
"""

from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .errors import WorkbookIntegrityError
from .formula_eval import ColumnMap, FormulaEvaluator, apply_template, col_to_index, index_to_col
from .model import (
    CellStyle,
    SheetModel,
    get_conditional_style,
    populate,
)
from .workbook import atomic_save, open_workbook


@dataclass(frozen=True)
class CellOp:
    """A single planned cell mutation."""

    sheet: str
    row: int
    col: int
    value: Any
    is_formula: bool = False
    style: CellStyle | None = None


# ── value normalisation ──────────────────────────────────────────────────────
def _py(value: Any) -> Any:
    """Convert a pandas/numpy value to a plain Python value openpyxl accepts."""
    if value is None:
        return None
    if isinstance(value, float) and pd.isna(value):
        return None
    try:
        if pd.isna(value):  # handles pd.NA / NaT for scalars
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(value, (pd.Timestamp,)):
        return value.to_pydatetime()
    if hasattr(value, "item"):  # numpy scalar
        try:
            return value.item()
        except (ValueError, AttributeError):
            return value
    if isinstance(value, (datetime, date)):
        return value
    return value


# ── column letter assignment ─────────────────────────────────────────────────
def _column_letters(model: SheetModel) -> list[tuple]:
    """Return [(column, col_index)] assigning positional letters where absent."""
    out = []
    pos = 0
    for col in model.columns:
        pos += 1
        if col.letter:
            from .formula_eval import col_to_index

            idx = col_to_index(col.letter)
        else:
            idx = pos
        out.append((col, idx))
    return out


# ── pure planning (data region only; testable without a workbook) ─────────────
def plan_data_ops(models: list[SheetModel]) -> list[CellOp]:
    """Plan header + data-row cell ops for each model. Pure; no workbook access.

    Region relocation and residual clearing need the live workbook and happen in
    :func:`render`.
    """
    ops: list[CellOp] = []
    for model in models:
        layout = model.layout
        cols = _column_letters(model)

        if layout.write_headers:
            hrow = max(layout.header_rows)
            for col, idx in cols:
                style = _compose(model.default_header_style, col.header_style)
                ops.append(CellOp(model.name, hrow, idx, col.header, style=style))

        frame = model.frame.reset_index(drop=True)
        for i in range(len(frame)):
            dest_row = layout.data_start_row + i
            row_map = frame.iloc[i].to_dict()
            for col, idx in cols:
                if col.writes_formula:
                    value = apply_template(col.formula.template, dest_row)
                    is_formula = str(value).startswith("=")
                else:
                    value = _py(row_map.get(col.key))
                    is_formula = False
                # Compose: sheet default → column style → conditional override.
                style = _compose(model.default_style, col.style)
                if col.conditional_style:
                    override = get_conditional_style(col.conditional_style)(row_map)
                    if override is not None:
                        style = _compose(style, override)
                ops.append(
                    CellOp(model.name, dest_row, idx, value, is_formula, style)
                )
    return ops


def _compose(base: CellStyle | None, override: CellStyle | None) -> CellStyle | None:
    """Merge ``override`` onto ``base``; either may be None."""
    if base is None:
        return override
    return base.merge(override)


# ── openpyxl style application ───────────────────────────────────────────────
def _argb(color: str) -> str:
    color = color.lstrip("#").upper()
    return color if len(color) == 8 else "FF" + color


_THIN = Side(style="thin")
_BOX = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_H_ALIGN = {"left", "center", "right", "general", "justify"}
_V_ALIGN = {"top", "center", "bottom"}


def _apply_style(cell, style: CellStyle | None) -> None:
    if style is None:
        return
    if style.number_format:
        cell.number_format = style.number_format
    if (
        style.bold
        or style.italic
        or style.font_size
        or style.font_color
        or style.font_name
    ):
        cell.font = Font(
            name=style.font_name,
            bold=style.bold,
            italic=style.italic,
            size=style.font_size,
            color=_argb(style.font_color) if style.font_color else None,
        )
    if style.fill:
        cell.fill = PatternFill(fill_type="solid", fgColor=_argb(style.fill))
    per_side = (style.border_top, style.border_bottom, style.border_left, style.border_right)
    if style.border or style.border_style or style.border_color or any(per_side):
        color = _argb(style.border_color) if style.border_color else None
        base = (
            Side(style=style.border_style or "thin", color=color)
            if (style.border or style.border_style)
            else None
        )

        def _side(per: str | None):
            return Side(style=per, color=color) if per else base

        cell.border = Border(
            left=_side(style.border_left),
            right=_side(style.border_right),
            top=_side(style.border_top),
            bottom=_side(style.border_bottom),
        )
    if style.align or style.valign or style.wrap:
        cell.alignment = Alignment(
            horizontal=style.align if style.align in _H_ALIGN else None,
            vertical=style.valign if style.valign in _V_ALIGN else None,
            wrap_text=style.wrap or None,
        )


# ── preserved-region capture (read-only pass) ────────────────────────────────
@dataclass(frozen=True)
class _CapturedCell:
    """A preserved-region cell snapshot: value + full styling."""

    col: int
    value: Any
    font: Any
    fill: Any
    border: Any
    alignment: Any
    number_format: str


def _needs_capture(models: list[SheetModel]) -> bool:
    # Any preserved region must be captured, because the data-region clear below
    # wipes styling as well as values and would otherwise erase the region.
    return any(m.layout.preserved_regions for m in models)


def _capture_regions(path: Path, models: list[SheetModel]) -> dict:
    """Read-only capture of preserved regions: per cell, value + full styling.

    Captured so a region can be relocated faithfully *after* the data region
    (now including its styling) is cleared. Values come from a ``data_only`` read
    so cross-sheet formula references are frozen to their computed values.
    """
    captured: dict = {}
    with open_workbook(path, read_only=False, data_only=True) as wb:
        for model in models:
            if model.name not in wb.sheetnames:
                continue
            ws = wb[model.name]
            sheet_caps = []
            for region in model.layout.preserved_regions:
                end = region.end_row or ws.max_row
                rows = []
                for r in range(region.start_row, end + 1):
                    cells = [
                        _CapturedCell(
                            col=c,
                            value=cell.value,
                            font=copy(cell.font),
                            fill=copy(cell.fill),
                            border=copy(cell.border),
                            alignment=copy(cell.alignment),
                            number_format=cell.number_format,
                        )
                        for c in range(1, ws.max_column + 1)
                        for cell in (ws.cell(r, c),)
                    ]
                    rows.append(cells)
                sheet_caps.append((region, rows))
            captured[model.name] = sheet_caps
    return captured


# ── main entry ───────────────────────────────────────────────────────────────
def render(
    path: str | Path,
    models: list[SheetModel],
    *,
    evaluator: FormulaEvaluator | None = None,
    colmap: ColumnMap | None = None,
    create_missing: bool = False,
    skip_populate: bool = False,
) -> Path:
    """Write ``models`` into the workbook at ``path`` in one pass.

    The workbook at ``path`` is used as the template (existing sheets and
    formatting are preserved). Set ``create_missing=True`` to allow creating a
    sheet a model targets but the template lacks.

    Unless ``skip_populate`` is set, each model is :func:`~.model.populate`d
    first so formula/computed columns hold authoritative values.
    """
    path = Path(path)
    if not skip_populate:
        for model in models:
            populate(model, evaluator=evaluator, colmap=colmap)

    captured = _capture_regions(path, models) if _needs_capture(models) else {}
    ops = plan_data_ops(models)

    template_exists = path.exists()
    if template_exists:
        cm = open_workbook(path, read_only=False, data_only=False)
        wb = cm.__enter__()
    else:
        wb = Workbook()  # the default "Sheet" is dropped after models add theirs
        cm = None

    try:
        names_before = set(wb.sheetnames)
        for model in models:
            if model.name not in wb.sheetnames:
                if not (create_missing or not template_exists):
                    raise WorkbookIntegrityError(
                        f"sheet {model.name!r} not found in template "
                        f"{path.name!r}; pass create_missing=True to add it"
                    )
                wb.create_sheet(model.name)
            ws = wb[model.name]
            _clear_and_relocate(ws, model, captured.get(model.name, []))

        # Apply planned data/header ops.
        for op in ops:
            cell = wb[op.sheet].cell(op.row, op.col)
            cell.value = op.value
            _apply_style(cell, op.style)

        # Fixed extra cells (outside the data/header region, e.g. a subtotal row).
        for model in models:
            for row, letter, value, style in model.extra_cells:
                cell = wb[model.name].cell(row, col_to_index(letter))
                cell.value = value
                _apply_style(cell, style)

        # Column widths.
        for model in models:
            for col, idx in _column_letters(model):
                if col.width:
                    wb[model.name].column_dimensions[index_to_col(idx)].width = col.width

        # Drop the default sheet if we created a fresh workbook with real sheets.
        if not template_exists and "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb["Sheet"]

        _integrity_guard(wb, names_before, models, template_exists)
        atomic_save(wb, path)
    finally:
        if cm is not None:
            cm.__exit__(None, None, None)
        else:
            wb.close()
    return path


def _last_nonempty_row(ws, start: int, ceiling: int, max_col: int) -> int:
    """Last row in [start, ceiling] with any non-empty cell; start-1 if none."""
    last = start - 1
    for r in range(start, ceiling + 1):
        if any(ws.cell(r, c).value is not None for c in range(1, max_col + 1)):
            last = r
    return last


def _clear_and_relocate(ws, model: SheetModel, sheet_caps: list) -> None:
    """Clear the old data region and relocate preserved regions by the row delta.

    The delta is based on the *actual* number of old data rows (scanned), not the
    distance to the preserved block — the gap between data and the block is
    intentional padding that must stay constant.
    """
    layout = model.layout
    start = layout.data_start_row
    max_col = ws.max_column or 1
    regions = layout.preserved_regions
    ceiling = (min(r.start_row for r in regions) - 1) if regions else (ws.max_row or start)

    old_last = _last_nonempty_row(ws, start, ceiling, max_col)
    old_len = max(0, old_last - start + 1)
    new_len = len(model.frame)
    delta = new_len - old_len
    # Note: a relocated region can never collide with the data region — old_len is
    # scanned only up to the region, so the region always lands at least one row
    # below the last data row. The padding between data and region is preserved.

    # Clear values AND styling from data_start to the end (data + old preserved
    # positions). Blanking styling too means emptied rows carry no stale fills or
    # borders; newly written rows are styled fresh from the model.
    for r in range(start, (ws.max_row or start) + 1):
        for c in range(1, max_col + 1):
            _blank_cell(ws.cell(r, c))

    # Rewrite preserved regions at shifted positions, restoring value + full style.
    for region, rows in sheet_caps:
        for r_offset, cells in enumerate(rows):
            new_row = region.start_row + delta + r_offset
            for cap in cells:
                cell = ws.cell(new_row, cap.col)
                cell.value = cap.value
                cell.font = cap.font
                cell.fill = cap.fill
                cell.border = cap.border
                cell.alignment = cap.alignment
                cell.number_format = cap.number_format


def _blank_cell(cell) -> None:
    """Reset a cell's value and all visible styling to openpyxl defaults."""
    cell.value = None
    cell.font = Font()
    cell.fill = PatternFill()
    cell.border = Border()
    cell.alignment = Alignment()
    cell.number_format = "General"


def _integrity_guard(
    wb, names_before: set[str], models: list[SheetModel], template_existed: bool
) -> None:
    names_after = set(wb.sheetnames)
    if not template_existed:
        # The fresh workbook's placeholder "Sheet" is not operator content.
        names_before = names_before - {"Sheet"}
        names_after = names_after - {"Sheet"}
    dropped = names_before - names_after
    if dropped:
        raise WorkbookIntegrityError(
            f"render would drop sheet(s) {sorted(dropped)} — preserve-by-default "
            f"was violated"
        )
